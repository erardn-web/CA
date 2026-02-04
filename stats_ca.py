import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime

# --- CONFIGURATION ---
ONGLET_SOURCE = "Prestation"
ONGLET_CIBLE = "stats 2026"
ANNEE = 2026

CODES_SEANCE = [
    "7311", "7301", "7340", "7601", "25.110", "1062",
    "1062-45", "3101", "7330", "7611", "7621",
    "privé", "Foyer de jour repas"
]

PHYSIOS = {
    997, 2171, 6620, 5787, 3646, 3933, 1613, 998,
    3309, 2248, 7271, 995, 1151, 5401, 3436
}

ERGOS = {
    7014: "Amir",
    6418: "Camille",
    5303: "Cindy",
    5783: "David",
    6911: "Younès",
    4516: "Roxane",
}

MASSEUSE = {3363: "Louise"}

LIGNES_CA = {
    "Louise": 10, "Roxane": 11, "Cindy": 12, "David": 13,
    "Younès": 14, "Amir": 15, "Camille": 16,
    "Physiothérapeutes": 18
}

LIGNES_SEANCES = {
    "Physiothérapeutes": 35, "Louise": 37, "Roxane": 38,
    "Cindy": 39, "David": 40, "Younès": 41,
    "Amir": 42, "Camille": 43
}

LIGNE_CA_PHYSIO_MENSUEL = 57
LIGNE_CA_ERGO_MENSUEL = 58

# --- LOGIQUE DE CALCUL ---

def charger_donnees(file_source):
    df = pd.read_excel(file_source, sheet_name=ONGLET_SOURCE)

    df["ID"] = df["Thérapeute"].astype(str).str.extract(r"\((\d+)\)")
    df = df.dropna(subset=["ID"])
    df["ID"] = df["ID"].astype(int)

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df[df["Date"].dt.year == ANNEE]
    df = df[df["Chiffre"] > 0]

    df["Code tarifaire"] = df["Code tarifaire"].astype(str)
    df["EstSéance"] = df["Code tarifaire"].isin(CODES_SEANCE)

    def categorie(id_):
        if id_ in PHYSIOS: return "Physiothérapeutes"
        if id_ in ERGOS: return ERGOS[id_]
        if id_ in MASSEUSE: return "Louise"
        return None

    df["Categorie"] = df["ID"].apply(categorie)
    df = df.dropna(subset=["Categorie"])
    df["Semaine"] = df["Date"].dt.isocalendar().week
    df["Mois"] = df["Date"].dt.month
    return df

def mettre_a_jour_excel(df, file_cible):
    # Charger le fichier cible en mémoire
    wb = openpyxl.load_workbook(file_cible)
    if ONGLET_CIBLE not in wb.sheetnames:
        st.error(f"L'onglet '{ONGLET_CIBLE}' est introuvable dans le fichier cible.")
        return None
    ws = wb[ONGLET_CIBLE]

    # 1. Calcul Hebdomadaire
    ca = df.groupby(["Categorie", "Semaine"])["Chiffre"].sum()
    seances = df[df["EstSéance"]].groupby(["Categorie", "Semaine"]).size()

    for (cat, sem), val in ca.items():
        if cat in LIGNES_CA and 1 <= sem <= 52:
            ws.cell(row=LIGNES_CA[cat], column=sem + 1).value = float(round(val))

    for (cat, sem), val in seances.items():
        if cat in LIGNES_SEANCES and 1 <= sem <= 52:
            ws.cell(row=LIGNES_SEANCES[cat], column=sem + 1).value = int(val)

    # 2. Calcul Mensuel
    df_physio = df[df["Categorie"].isin(["Physiothérapeutes", "Louise"])]
    df_ergo = df[df["Categorie"].isin(ERGOS.values())]

    for mois, val in df_physio.groupby("Mois")["Chiffre"].sum().items():
        ws.cell(row=LIGNE_CA_PHYSIO_MENSUEL, column=mois + 2).value = float(round(val))

    for mois, val in df_ergo.groupby("Mois")["Chiffre"].sum().items():
        ws.cell(row=LIGNE_CA_ERGO_MENSUEL, column=mois + 2).value = float(round(val))

    # Sauvegarde dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- INTERFACE WEB STREAMLIT ---

st.set_page_config(page_title="Calcul CA Thérapeutes", layout="centered")
st.title("📊 Calculateur de Chiffre d'Affaires 2026")
st.markdown("---")

st.info("👋 Bienvenue. Importez les deux fichiers pour générer vos statistiques mises à jour.")

col1, col2 = st.columns(2)
with col1:
    file_source = st.file_uploader("📂 Fichier Source (Prestation)", type=["xlsx"])
with col2:
    file_cible = st.file_uploader("🎯 Fichier Cible (Stats 2026)", type=["xlsx"])

if file_source and file_cible:
    if st.button("🚀 Lancer tous les calculs", type="primary", use_container_width=True):
        with st.spinner("Analyse et injection des données en cours..."):
            try:
                # Execution
                df_traite = charger_donnees(file_source)
                result_buffer = mettre_a_jour_excel(df_traite, file_cible)
                
                if result_buffer:
                    st.success("✅ Calculs terminés avec succès !")
                    
                    # Bouton de téléchargement
                    date_str = datetime.now().strftime("%d-%m_%Hh%M")
                    st.download_button(
                        label="📥 Télécharger le fichier Stats mis à jour",
                        data=result_buffer,
                        file_name=f"Statistiques_2026_Maj_{date_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            except Exception as e:
                st.error(f"Une erreur est survenue lors du calcul : {e}")

st.markdown("---")
st.caption("Note : Ce script n'enregistre aucune donnée sur le serveur. Vos fichiers sont traités en mémoire vive.")
